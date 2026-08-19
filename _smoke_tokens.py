import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import backup_all as B

expected = {
    "出库记录": "Uuu0feVP0lbzEydQaLQcElJ9nVg",
    "库存": "LaBjfwhEMlY0Hcdl0smczvcVnEd",
    "销售订单": "JCm8fVnNalN1UGd2nmxczS6SnMd",
    "销售订单明细汇总": "P30pfKZFQlzDbBduojfc6DDHnFe",
    "采购订单": "Mrf3fb7fEl5CVqdwgwqcWn8qnmd",
    "采购订单明细汇总": "ElTnfu6CvlVfbNdxndycPJUGnTd",
    "收付确认": "APyCfFjQ4lO7PxdFblucyWeAnXe",
    "应收结算": "L5TufudKAluIrBduRtIcwACjnXd",
    "费用管理": "LZxofXocflINVGdngtacqJWvnbe",
    "开票明细": "JvZMfAgnulJUWBdHVLpclXTTn4f",
}
ok = True
for t in B.TASKS:
    want = expected[t["biz"]]
    got = t.get("folder_token", "")
    if got != want:
        ok = False
        print(f"[MISMATCH] {t['biz']}: want={want} got={got}")
    if got == "" or "folder_env" in t:
        ok = False
        print(f"[BAD] {t['biz']}: empty token or still has folder_env key")
print("TASKS folder_token 全部匹配:", ok, f"({len(B.TASKS)} entries)")

b = B.records_to_xlsx_bytes([
    {"id": "1", "单号": "X001", "金额": "1234.56", "日期": "2026-08-17"},
    {"id": "2", "单号": "X002", "金额": "0", "日期": "2026-08-18"},
], summary_pairs=[("合计金额", "1234.56"), ("单数", "2")])
from openpyxl import load_workbook
import io
wb = load_workbook(io.BytesIO(b), data_only=True)
print("sheets:", wb.sheetnames)
ws = wb["数据"]
a2 = ws["A2"].value; c2 = ws["C2"].value; c3 = ws["C3"].value
print(f"  数据 A2: {a2!r} ({type(a2).__name__})  应为 int 1")
print(f"  数据 C2: {c2!r} ({type(c2).__name__})  应为 float 1234.56")
print(f"  数据 C3: {c3!r} ({type(c3).__name__})  应为 int 0")
print(f"  冻结: {ws.freeze_panes}  应为 A2")
assert isinstance(a2, int) and a2 == 1, f"A2 类型错: {type(a2).__name__}"
assert isinstance(c2, float) and abs(c2 - 1234.56) < 1e-9, f"C2 类型或值错"
assert isinstance(c3, int) and c3 == 0, f"C3 类型或值错"
ws2 = wb["汇总"]
hj = ws2["B2"].value
print(f"  汇总 B2(合计金额): {hj!r} ({type(hj).__name__})  应为 float 1234.56")
assert isinstance(hj, float), f"汇总B2 应为 float: {type(hj).__name__}"
print("\n所有冒烟检查通过 ✅")
