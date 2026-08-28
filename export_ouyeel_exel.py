#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
把飞书多维表格 tblDFiiwWkJ5tZp3 中「挂士禾」「挂砚启」两视图的记录，
按欧冶模板字段映射生成两个 Excel 文件，再通过飞书 IM 私聊发给指定用户，
并附上总结汇报文本。

环境变量：
  FEISHU_APP_ID / FEISHU_APP_SECRET   (必需)
  BITABLE_APP_TOKEN                   (可选，默认 IRCzb5XVganwbUsz9NKcaC0SnRA)
  FEISHU_UNION_IDS                    (可选，逗号分隔；默认同 export_lindiao 的两个 union_id)
"""
import os
import sys
import io
import json
import time
import datetime
import requests

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

APP_TOKEN_DEFAULT = "IRCzb5XVganwbUsz9NKcaC0SnRA"
TABLE_ID = "tblDFiiwWkJ5tZp3"
FEISHU_OPEN_BASE = "https://open.feishu.cn/open-apis"

# 视图名: view_id (注意：飞书表中的实际视图名是"挂士禾"/"挂砚启"，与用户说的「挂-士禾」「挂-砚启」略有不同)
VIEWS = {
    "挂-士禾": "vewV7IQKIn",
    "挂-砚启": "vewiMib61q",
}

# 收件人默认列表（英文逗号分隔）：
#   on_ 开头 = 个人 union_id（私聊）
#   oc_ 开头 = 群 chat_id（群聊），脚本按前缀自动识别 receive_id_type
DEFAULT_UNION_IDS = (
    "on_93da40c6314edbfa2dc3e031ef405389,"
    "on_b09bcbf3e74f5d423900aa9b2f00eb63,"
    "oc_334f8c12e73592af76dccb5b34ccfa5f,"
    "oc_d22e1f9c8cd0a5a3aa2b2625e2a8f155"
)

# ------------------------------------------------------------
# 欧冶模板完整列头（共 29 列）
#   第 1 行：必填项(点击进入使用说明)  × 13 列  +  选填项 × 1 列(质量缺陷必填区) + 选填项 × 15 列
#   第 2 行：各列具体字段名
# ------------------------------------------------------------
TEMPLATE_HEADERS = [
    # =========== 前 13 列：必填项区域 ===========
    "品种\n点击选择",          # 1
    "产地\n点击选择",          # 2
    "仓库\n点击选择",          # 3
    "牌号",                    # 4
    "规格\n（厚mm*宽mm*长mm）", # 5
    "重量\n(吨）",             # 6
    "捆包号",                  # 7
    "质量等级",                # 8
    "挂牌价格\n（含税）",      # 9
    "资源区",                  # 10
    "是否有质保信息",          # 11
    "是否原卷原标签",          # 12
    "质量缺陷",                # 13  （必填项区内单独一列）
    # =========== 第 14 列：是否后结算（仍属于必填项区之前列位置，按用户样例此处也在"必填"）========
    "是否\n后结算",            # 14
    # =========== 从 15 列起：15 个选填项 ===========
    "参考牌号",                # 15
    "钢厂资源号",              # 16
    "技术标准",                # 17
    "最低挂牌价",              # 18
    "表面处理\n点击选择",      # 19
    "入库时间",                # 20
    "出厂日期",                # 21
    "特殊说明",                # 22   (原 14 列位置，现在挪到第 22 列)
    "镀层种类",                # 23
    "面漆种类",                # 24
    "颜色",                    # 25
    "规格详情",                # 26
    "QZA判断",                 # 27
    "库位",                    # 28
    "包装方式",                # 29
]

# 第 1 行表头：每一段的 (开始列, 结束列, 显示文字, 底色)
# 列号 1-based。根据用户给出的原始表格：
#   - 前 13 列标题都是「必填项(点击进入使用说明）」
#   - 从第 14 列("是否\n后结算")开始，用户模板原文中第 14 列顶格未写必填/选填字样，
#     但紧邻右边 15~29 列顶部都写了「选填项」。我们按原文处理：
#     A1..M1 (1..13) = "必填项(点击进入使用说明）"
#     N1      (14)   = ""  (用户给的是空白在"是否后结算"上方)
#     O1..AC1(15..29) = "选填项"
# 但为了视觉上更匹配用户给的「选填项 × 16 列」的 16 列宽度，
# 我们使用下面这个精确的段切分：
HEADER_ROW_1_SPANS = [
    # (start_col, end_col, label, fill)
    (1,  13, "必填项(点击进入使用说明）", "FFE699"),  # 浅橙：必填区
    (14, 14, "",                          "FFFFFF"),
    (15, 29, "选填项",                     "DDEBF7"),  # 浅蓝：选填区
]

# ------------------------------------------------------------
# 源表字段 → 模板字段 的映射。
# 若源表不存在该字段（例如"技术标准""出厂日期"等），则映射为 None，
# 写入空字符串（用户样例也可能为空）。
# ------------------------------------------------------------
FIELD_MAPPING = {
    # -------- 前 13 列（必填项区） --------
    "品种\n点击选择":            "品种",
    "产地\n点击选择":            "产地",
    "仓库\n点击选择":            "仓库",
    "牌号":                      "牌号",
    "规格\n（厚mm*宽mm*长mm）":   "规格",
    "重量\n(吨）":               "重量",
    "捆包号":                    "捆包号",
    "质量等级":                  "质量等级",
    "挂牌价格\n（含税）":        "挂牌价格",
    "资源区":                    "资源区",
    "是否有质保信息":            "是否有质保信息",
    "是否原卷原标签":            "是否原卷标签",     # 源字段名少一个"原"字
    "质量缺陷":                  "质量缺陷",
    # -------- 第 14 列 --------
    "是否\n后结算":              "是否后结算",
    # -------- 15~29 列（选填区） --------
    "参考牌号":                  "参考牌号",
    "钢厂资源号":                "钢厂资源号",
    "技术标准":                  None,              # 源表暂无
    "最低挂牌价":                None,              # 源表暂无；但如果源表挂牌价相同则可能一样，留空
    "表面处理\n点击选择":        None,              # 源表暂无
    "入库时间":                  "入库日期",        # 源表有"入库日期" datetime，直接用
    "出厂日期":                  None,              # 源表暂无
    "特殊说明":                  "特殊说明",
    "镀层种类":                  "锌层",            # 源表"锌层"字段近似对应欧冶"镀层种类"
    "面漆种类":                  "涂料",            # 源表"涂料"近似对应欧冶"面漆种类"
    "颜色":                      "颜色",
    "规格详情":                  "结构",            # 源表"结构"近似（若是镀锡"结构"=D/FC 等则近似规格详情）
    "QZA判断":                   None,
    "库位":                      "库位号",          # 源表有"库位号"
    "包装方式":                  None,
}

# 源表 datetime(ms) → 本地 YYYY-MM-DD 字符串
def _ms_to_date_str(ms_val):
    try:
        n = int(ms_val)
    except Exception:
        return ""
    if not n:
        return ""
    dt = datetime.datetime.utcfromtimestamp(n / 1000) + datetime.timedelta(hours=8)
    return dt.strftime("%Y-%m-%d")


# =========================================================================
# 通用辅助
# =========================================================================
def env(key, default=None):
    return os.environ.get(key, default)


def feishu_tenant_access_token(app_id, app_secret) -> str:
    url = f"{FEISHU_OPEN_BASE}/auth/v3/tenant_access_token/internal"
    r = requests.post(url, json={"app_id": app_id, "app_secret": app_secret}, timeout=15)
    data = r.json()
    if data.get("code") != 0:
        raise RuntimeError(f"换 tenant_access_token 失败: {data}")
    token = data.get("tenant_access_token")
    if not token:
        raise RuntimeError(f"响应中无 tenant_access_token: {data}")
    print(f"[飞书] tenant_access_token 获取成功 len={len(token)}")
    return token


# =========================================================================
# 多维表格: 按 view_id 全量分页拉取
# =========================================================================
def bitable_list_records_by_view(token: str, app_token: str, table_id: str,
                                  view_id: str) -> list[dict]:
    url = f"{FEISHU_OPEN_BASE}/bitable/v1/apps/{app_token}/tables/{table_id}/records"
    headers = {"Authorization": f"Bearer {token}"}
    records: list[dict] = []
    page_token = ""
    page = 1
    while True:
        params = {"page_size": 500, "view_id": view_id}
        if page_token:
            params["page_token"] = page_token
        r = requests.get(url, headers=headers, params=params, timeout=60)
        data = r.json()
        if data.get("code") != 0:
            raise RuntimeError(f"bitable 拉取 records 失败: code={data.get('code')}, "
                               f"msg={data.get('msg')}, raw={data}")
        items = (data.get("data") or {}).get("items") or []
        records.extend(items)
        print(f"  [bitable] view={view_id} page {page}: +{len(items)} 累计 {len(records)}")
        if not (data.get("data") or {}).get("has_more"):
            break
        page_token = (data.get("data") or {}).get("page_token") or ""
        page += 1
        time.sleep(0.1)
    return records


# =========================================================================
# 源字段值扁平化：text 富文本 / 数组 → 字符串；数值保留
# =========================================================================
def flatten(v):
    """把 bitable 返回的字段值压成扁平字符串或数值。"""
    if v is None:
        return ""
    # 数值 / 字符串
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, str):
        return v
    # list：常见 [{"text": ...}] 或 [number]
    if isinstance(v, list):
        parts = []
        for x in v:
            if isinstance(x, dict):
                # 富文本段优先 text
                if "text" in x:
                    parts.append(str(x["text"]))
                else:
                    parts.append(json.dumps(x, ensure_ascii=False))
            else:
                parts.append(str(x))
        s = ",".join(p for p in parts if p != "")
        # 全是数值列表的情况（比如 重量=[13.69]），尝试返回单个数值
        if len(v) == 1 and isinstance(v[0], (int, float)):
            return v[0]
        return s
    if isinstance(v, dict):
        return json.dumps(v, ensure_ascii=False)
    return str(v)


# 数值/日期列的 header（用于 number_format 判定）
_NUMBER_HEADERS = {"重量\n(吨）": "0.000",
                   "挂牌价格\n（含税）": "¥#,##0",
                   "最低挂牌价": "¥#,##0"}
_TEXT_LEFT_HEADERS = {"捆包号", "钢厂资源号", "规格\n（厚mm*宽mm*长mm）", "规格详情"}


# =========================================================================
# 生成 Excel (bytes) —— 两行表头（第 1 行：必填/选填段合并；第 2 行：具体字段名）
# =========================================================================
def make_excel_bytes(rows: list[dict]) -> bytes:
    """
    rows: 已按 TEMPLATE_HEADERS 为 key 的字典列表
    """
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "导入数据"

    header_font   = Font(bold=True, name="宋体", size=11)
    header1_font  = Font(bold=True, name="宋体", size=12, color="C00000")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    right  = Alignment(horizontal="right",  vertical="center", wrap_text=True)
    thin_side = Side(style="thin", color="7F7F7F")
    border  = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header2_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    white_fill   = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    N = len(TEMPLATE_HEADERS)  # 29

    # -------- 第 1 行：分段合并 + 标题 --------
    for (start, end, label, color_hex) in HEADER_ROW_1_SPANS:
        span_fill = PatternFill(start_color=color_hex, end_color=color_hex,
                                fill_type="solid")
        if start != end:
            ws.merge_cells(start_row=1, start_column=start,
                           end_row=1,   end_column=end)
        c = ws.cell(row=1, column=start, value=label)
        c.font = header1_font
        c.alignment = center
        # 给段内每一格加上边框 + 底色
        for col_idx in range(start, end + 1):
            cc = ws.cell(row=1, column=col_idx)
            cc.fill = span_fill
            cc.border = border
    ws.row_dimensions[1].height = 26

    # -------- 第 2 行：具体列头 --------
    for col_idx, h in enumerate(TEMPLATE_HEADERS, 1):
        c = ws.cell(row=2, column=col_idx, value=h)
        c.font = header_font
        c.fill = header2_fill
        c.alignment = center
        c.border = border
    ws.row_dimensions[2].height = 40

    # -------- 数据行（从第 3 行开始） --------
    for r in rows:
        line = [r.get(h, "") for h in TEMPLATE_HEADERS]
        ws.append(line)
        data_row = ws.max_row
        for col_idx in range(1, N + 1):
            header = TEMPLATE_HEADERS[col_idx - 1]
            c = ws.cell(row=data_row, column=col_idx)
            c.border = border
            if header in _NUMBER_HEADERS and isinstance(c.value, (int, float)):
                c.number_format = _NUMBER_HEADERS[header]
                c.alignment = right
            elif header in _TEXT_LEFT_HEADERS:
                c.alignment = left
                if header == "捆包号":
                    c.number_format = "@"  # 文本，避免长数字科学计数
            else:
                c.alignment = left

    # -------- 列宽（29 列） --------
    widths = [
        12, 12, 38, 16, 26, 10, 22, 12, 14, 10,   # 1~10
        16, 18, 12, 10, 16, 18, 14, 12, 12, 14,   # 11~20
        14, 46, 12, 12, 12, 18, 10, 12, 12,       # 21~29
    ]
    for i, w in enumerate(widths[:N], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 冻结前两行 + 首列
    ws.freeze_panes = "B3"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =========================================================================
# 飞书上传文件到 IM 资源 (获取 file_key 供发文件消息使用)
# =========================================================================
def feishu_im_upload_file(token: str, file_bytes: bytes, filename: str,
                          file_type: str = "xlsx") -> str:
    """
    通过 /im/v1/files 接口把文件上传到飞书 IM，返回 file_key。
    file_type 参考飞书文档：mp4/pdfs/doc/xls/ppt/xlsx/docx/pptx 等
    """
    url = f"{FEISHU_OPEN_BASE}/im/v1/files"
    headers = {"Authorization": f"Bearer {token}"}
    data = {"file_type": file_type, "file_name": filename}
    files = {"file": (filename, file_bytes,
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = requests.post(url, headers=headers, data=data, files=files, timeout=120)
    resp = r.json()
    if resp.get("code") != 0:
        raise RuntimeError(f"IM 上传文件失败 {filename}: code={resp.get('code')}, "
                           f"msg={resp.get('msg')}, raw={resp}")
    fk = (resp.get("data") or {}).get("file_key") or ""
    if not fk:
        raise RuntimeError(f"IM 上传文件返回中无 file_key: {resp}")
    print(f"[飞书IM] 上传文件 {filename} 成功, file_key={fk[:10]}... 大小 {len(file_bytes)}B")
    return fk


# =========================================================================
# 飞书 IM 发消息（文本 / 文件）
# =========================================================================
def receive_id_type_of(rid: str) -> str:
    """按前缀自动识别：oc_ 开头 = 群 chat_id，其余按个人 union_id 处理。"""
    return "chat_id" if rid.startswith("oc_") else "union_id"


def feishu_send_im_text(token: str, receive_id: str, text: str,
                        receive_id_type: str = None):
    receive_id_type = receive_id_type or receive_id_type_of(receive_id)
    url = f"{FEISHU_OPEN_BASE}/im/v1/messages?receive_id_type={receive_id_type}"
    headers = {"Authorization": f"Bearer {token}"}
    payload = {
        "receive_id": receive_id,
        "msg_type": "text",
        "content": json.dumps({"text": text}, ensure_ascii=False),
    }
    r = requests.post(url, headers=headers, json=payload, timeout=30)
    data = r.json()
    if data.get("code") != 0:
        raise RuntimeError(f"IM 发文本失败 code={data.get('code')}, msg={data.get('msg')}, raw={data}")
    print(f"[飞书IM] 文本 -> {receive_id[:6]}... 成功")


def feishu_send_im_file(token: str, receive_id: str, file_key: str,
                        receive_id_type: str = None):
    receive_id_type = receive_id_type or receive_id_type_of(receive_id)
    url = f"{FEISHU_OPEN_BASE}/im/v1/messages?receive_id_type={receive_id_type}"
    headers = {"Authorization": f"Bearer {token}"}
    payload = {
        "receive_id": receive_id,
        "msg_type": "file",
        "content": json.dumps({"file_key": file_key}, ensure_ascii=False),
    }
    r = requests.post(url, headers=headers, json=payload, timeout=30)
    data = r.json()
    if data.get("code") != 0:
        raise RuntimeError(f"IM 发文件失败 code={data.get('code')}, msg={data.get('msg')}, raw={data}")
    print(f"[飞书IM] 文件 file_key={file_key[:10]}... -> {receive_id[:6]}... 成功")


# =========================================================================
# 主流程
# =========================================================================
def main() -> int:
    # 1) 取配置
    fs_app_id     = env("FEISHU_APP_ID")
    fs_app_secret = env("FEISHU_APP_SECRET")
    if not fs_app_id or not fs_app_secret:
        print("[错误] 缺少 FEISHU_APP_ID / FEISHU_APP_SECRET 环境变量")
        return 2

    bt_app_token = env("BITABLE_APP_TOKEN") or APP_TOKEN_DEFAULT
    # 收件人：优先用 FEISHU_UNION_ID_OUYEE（临时覆盖），否则用默认
    union_ids_raw = env("FEISHU_UNION_ID_OUYEE") or DEFAULT_UNION_IDS
    union_ids = [u.strip() for u in union_ids_raw.split(",") if u.strip()]
    if not union_ids:
        print("[错误] 未配置任何收件人（FEISHU_UNION_ID_OUYEE / 默认值）")
        return 2

    print(f"[配置] app_id={fs_app_id[:6]}... bitable_app_token={bt_app_token[:6]}...")
    labeled = [f"{u[:8]}...({'群' if u.startswith('oc_') else '个人'})" for u in union_ids]
    print(f"[配置] 收件人 {len(union_ids)} 个: " + ", ".join(labeled))

    # 2) 换 token
    fs_token = feishu_tenant_access_token(fs_app_id, fs_app_secret)

    # 3) 拉两个视图数据 → 映射 → 造 Excel bytes
    outputs: list[dict] = []   # [{name, bytes, rows, total_weight, ...}]
    # 需要把"毫秒时间戳 → YYYY-MM-DD"的列
    _DATE_COLUMNS = {"入库时间", "出厂日期"}

    for view_display, view_id in VIEWS.items():
        print(f"\n===== 读取视图 {view_display} (view_id={view_id}) =====")
        raw = bitable_list_records_by_view(fs_token, bt_app_token, TABLE_ID, view_id)
        print(f"[读取] {view_display} 共 {len(raw)} 条记录")

        mapped_rows: list[dict] = []
        total_weight = 0.0
        price_total  = 0.0
        missing_cols: dict[str, int] = {}

        for item in raw:
            f = item.get("fields") or {}
            row = {}
            for col_header in TEMPLATE_HEADERS:
                src_field = FIELD_MAPPING.get(col_header)
                if src_field is None:
                    # 源表没有这一列：留空
                    raw_val = ""
                else:
                    raw_val = flatten(f.get(src_field))
                    # 日期列：如果 flatten 出来是数字(ms 时间戳)，转 YYYY-MM-DD
                    if col_header in _DATE_COLUMNS and isinstance(raw_val, (int, float)):
                        raw_val = _ms_to_date_str(int(raw_val))
                row[col_header] = raw_val
                if raw_val in (None, ""):
                    missing_cols[col_header] = missing_cols.get(col_header, 0) + 1
            # 统计
            w = row["重量\n(吨）"]
            if isinstance(w, (int, float)):
                total_weight += float(w)
            p = row["挂牌价格\n（含税）"]
            if isinstance(p, (int, float)):
                price_total += float(p)
            mapped_rows.append(row)

        # 文件名：挂-士禾_YYMMDD_HHMMSS.xlsx
        ts = datetime.datetime.now().strftime("%y%m%d_%H%M%S")
        safe_name = view_display.replace("/", "-").replace("\\", "-")
        filename = f"{safe_name}_{ts}.xlsx"

        xlsx_bytes = make_excel_bytes(mapped_rows)
        # 空值最多的前 8 列（用于汇报）
        empty_top = sorted(missing_cols.items(), key=lambda x: -x[1])[:8]
        outputs.append({
            "view_name": view_display,
            "view_id": view_id,
            "filename": filename,
            "bytes": xlsx_bytes,
            "rows": len(mapped_rows),
            "total_weight": round(total_weight, 3),
            "avg_price": (round(price_total / len(mapped_rows), 2)
                          if mapped_rows else 0),
            "empty_top": empty_top,
        })
        print(f"[生成] {filename}  {len(xlsx_bytes)/1024:.1f} KB  "
              f"{len(mapped_rows)} 行, 总重 {total_weight:.3f} 吨")
        # 打一下空值前 5 列的概览，便于用户看哪些选填字段本来就没有数据
        if empty_top:
            print(f"  [空值较多列(前5)] " +
                  ", ".join(f"{k!r}空{v}/{len(mapped_rows)}"
                            for k, v in empty_top[:5]))

    # 4) 上传到 IM 资源得到 file_key
    for o in outputs:
        o["file_key"] = feishu_im_upload_file(fs_token, o["bytes"], o["filename"])

    # 5) 生成总结汇报文本（精简：只保留用户关心的信息）
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    summary_lines = [
        "【欧冶挂牌资源导出 · 总结汇报】",
        f"执行时间：{now}",
        "",
        "文件清单：",
    ]
    for idx, o in enumerate(outputs, 1):
        summary_lines.append(
            f"  {idx}) 文件名：{o['filename']}\n"
            f"     视图：{o['view_name']}\n"
            f"     记录数：{o['rows']} 条\n"
            f"     总重量：{o['total_weight']:.3f} 吨\n"
            f"     平均挂牌价：¥{o['avg_price']:.0f}（含税，简单平均）\n"
            f"     文件大小：{len(o['bytes'])/1024:.1f} KB"
        )
    # 空值前 5 列小结（便于人工核对哪些选填字段源表没填）
    for o in outputs:
        if o["empty_top"]:
            summary_lines.append(
                f"\n{o['view_name']} 空值较多列(前5)：" +
                "、".join(f"{k}({v}/{o['rows']})" for k, v in o["empty_top"][:5])
            )
    summary_text = "\n".join(summary_lines)

    # 6) 对每个收件人循环发送：先 2 个文件，再总结文本
    for uid in union_ids:
        print(f"\n===== 发送给 {uid[:8]}... =====")
        try:
            for o in outputs:
                feishu_send_im_file(fs_token, uid, o["file_key"])
                time.sleep(0.2)
            feishu_send_im_text(fs_token, uid, summary_text)
        except Exception as e:
            print(f"[警告] 发送给 {uid} 失败: {e}")
            # 其他收件人继续

    print("\n全部完成 ✓")
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception as exc:
        import traceback
        traceback.print_exc()
        sys.exit(1)
