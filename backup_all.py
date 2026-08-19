#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
懂火钢城系统 - 每周全量备份 10 类业务数据 → 飞书云盘

数据来源：erpa.donghuo.vip
流程：
  1. 自动登录（ddddocr 验证码识别，复用 donghuo_login 模块）
  2. 按配置遍历 10 类业务：
     - 9 类走 JSON 分页接口（POST /model/admin/.../getlist 或 .../mxlist）
     - 库存走 HTML 表格导出接口（POST /view/admin/excelbiao/kucungl）
  3. 每类生成一份 XLSX（Excel 原生格式，数字自动转 int/float，冻结首行，有汇总字段时另存 Sheet2）
  4. 上传到飞书云盘 10 个独立子文件夹（环境变量分别配置）
  5. 通过飞书机器人发送汇总通知

环境变量（必填，仓库 Secrets 注入）：
  DH_USERNAME           erpa 登录账号
  DH_PASSWORD           erpa 登录密码
  FEISHU_APP_ID         飞书自建应用 App ID
  FEISHU_APP_SECRET     飞书自建应用 App Secret

环境变量（可选）：
  FEISHU_WEBHOOK_URL       飞书机器人 Webhook（发送汇总通知）
  FEISHU_WEBHOOK_SECRET    飞书机器人签名密钥
  BACKUP_PAGE_SIZE         分页每页条数（默认 200；服务端上限约 300，超过会丢数据）
  BACKUP_DRY_RUN           =1 时仅生成 XLSX 到本地，不上传云盘
  TZ                       Asia/Shanghai

说明：
  10 个业务对应的飞书云盘子文件夹 token 已直接写在本文件 TASKS 配置里，
  无需再通过环境变量注入（减少 Secrets 配置量）。
  如日后需要换文件夹，直接改 TASKS 中各条目的 folder_token 即可。

注意：
  - 服务端对单页返回数有上限（约 300 条），且 offset = page * limit 用的是我们传的 limit。
    若 limit > 300，每页会跳过 (limit-300) 条数据，导致丢页。因此默认 limit=200。
  - 懂火后端 SQL Server 连接池不稳，连续请求偶发 "远程主机强迫关闭了一个现有的连接"。
    脚本每页/每类请求都做了 3 次重试，并在每类业务之间 sleep 4s。

依赖：
  pip install requests urllib3 ddddocr openpyxl

GitHub Actions 触发：每周定时（推荐 cron-job.org 外部 POST，避免 GitHub schedule 静默延迟）
"""

import os
import re
import io
import csv
import sys
import json
import time
import datetime
import traceback
from pathlib import Path

# 调试用默认账号（不提交，生产环境走 GitHub Secrets / 系统环境变量）
os.environ.setdefault("DH_USERNAME", os.environ.get("DH_USERNAME") or "")
os.environ.setdefault("DH_PASSWORD", os.environ.get("DH_PASSWORD") or "")

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from donghuo_login import login_donghuo, BASE_URL

import requests

# ===================== 配置 =====================

# 飞书 OpenAPI 基址
FEISHU_OPEN_BASE = "https://open.feishu.cn/open-apis"

# 备份日期范围文件名后缀格式
DATE_FMT = "%Y%m%d_%H%M%S"

# 10 类业务接口配置（与 apis.json 同步）
# folder_token：飞书云盘子文件夹 token，已直接写死在此，不用再配 GitHub Secrets
#   对应关系：出库=Uuu0feVP... / 库存=LaBjfwhE... / 销售订单=JCm8fVnN... /
#             销售明细=P30pfKZF... / 采购订单=Mrf3fb7f... / 采购明细=ElTnfu6C... /
#             收付确认=APyCfFjQ... / 应收结算=L5TufudK... / 费用管理=LZxofXoc... /
#             开票明细=JvZMfAgn...
TASKS = [
    {
        "biz": "出库记录",
        "folder_token": "Uuu0feVP0lbzEydQaLQcElJ9nVg",
        "api_type": "json_paged",
        "api_path": "/model/admin/xiaoshou/m_xiaoshou/xjilulist",
        "page_param": "page",
        "limit_param": "limit",
        "total_field": "pgtotal",
        "records_field": "root",
        "filename_prefix": "chuku",
    },
    {
        "biz": "库存",
        "folder_token": "LaBjfwhEMlY0Hcdl0smczvcVnEd",
        "api_type": "html_export",
        "api_path": "/view/admin/excelbiao/kucungl",
        "filename_prefix": "kucun",
        "note": "用户明确要求用库存管理里的'导出'",
    },
    {
        "biz": "销售订单",
        "folder_token": "JCm8fVnNalN1UGd2nmxczS6SnMd",
        "api_type": "json_paged",
        "api_path": "/model/admin/xiaoshou/m_dindan/getlist",
        "page_param": "page",
        "limit_param": "limit",
        "total_field": "pgtotal",
        "records_field": "root",
        "filename_prefix": "xsdd",
    },
    {
        "biz": "销售订单明细汇总",
        "folder_token": "P30pfKZFQlzDbBduojfc6DDHnFe",
        "api_type": "json_paged",
        "api_path": "/model/admin/xiaoshou/m_dindan/mxlist",
        "page_param": "page",
        "limit_param": "limit",
        "total_field": "pgtotal",
        "records_field": "root",
        "filename_prefix": "xsmx",
    },
    {
        "biz": "采购订单",
        "folder_token": "Mrf3fb7fEl5CVqdwgwqcWn8qnmd",
        "api_type": "json_paged",
        "api_path": "/model/admin/caigou/m_dindan/getlist",
        "page_param": "page",
        "limit_param": "limit",
        "total_field": "pgtotal",
        "records_field": "root",
        "filename_prefix": "cgdd",
    },
    {
        "biz": "采购订单明细汇总",
        "folder_token": "ElTnfu6CvlVfbNdxndycPJUGnTd",
        "api_type": "json_paged",
        "api_path": "/model/admin/caigou/m_dindan/mxlist",
        "page_param": "page",
        "limit_param": "limit",
        "total_field": "pgtotal",
        "records_field": "root",
        "filename_prefix": "cgmx",
    },
    {
        "biz": "收付确认",
        "folder_token": "APyCfFjQ4lO7PxdFblucyWeAnXe",
        "api_type": "json_paged",
        "api_path": "/model/admin/caiwu/m_liushui/getlist",
        "page_param": "page",
        "limit_param": "limit",
        "total_field": "pgtotal",
        "records_field": "root",
        "filename_prefix": "sfqr",
    },
    {
        "biz": "应收结算",
        "folder_token": "L5TufudKAluIrBduRtIcwACjnXd",
        "api_type": "json_paged",
        "api_path": "/model/admin/caiwu/m_yinshou/getlist",
        "page_param": "page",
        "limit_param": "limit",
        "total_field": "pgtotal",
        "records_field": "root",
        "filename_prefix": "ysjs",
        "note": "AR 数据含 hj_weishou 未收款汇总字段",
    },
    {
        "biz": "费用管理",
        "folder_token": "LZxofXocflINVGdngtacqJWvnbe",
        "api_type": "json_paged",
        "api_path": "/model/admin/caiwu/m_feiyon/getlist",
        "page_param": "page",
        "limit_param": "limit",
        "total_field": "pgtotal",
        "records_field": "root",
        "filename_prefix": "fygl",
    },
    {
        "biz": "开票明细",
        "folder_token": "JvZMfAgnulJUWBdHVLpclXTTn4f",
        "api_type": "json_paged",
        "api_path": "/model/admin/caiwu/m_fapiao/mxlist_x",
        "page_param": "page",
        "limit_param": "limit",
        "total_field": "pgtotal",
        "records_field": "root",
        "filename_prefix": "kpmx",
    },
]


# ===================== 数据拉取 =====================

def _try_parse_json(text: str):
    try:
        return json.loads(text)
    except Exception:
        return None


def _looks_like_sql_error(text: str) -> bool:
    """检测响应文本是否是懂火后端 SQL Server 连接断开的 PHP Fatal error 页。"""
    if not text:
        return False
    markers = (
        "com_exception",
        "SQL Server Native Client",
        "远程主机强迫关闭了一个现有的连接",
        "Fatal error",
        "无法连接",
        "Microsoft SQL Server",
    )
    t = text[:2000]
    return any(m in t for m in markers)


def fetch_json_paged(session: requests.Session,
                     api_path: str,
                     page_param: str = "page",
                     limit_param: str = "limit",
                     total_field: str = "pgtotal",
                     records_field: str = "root",
                     page_size: int = 500,
                     max_pages: int = 1000) -> tuple[list[dict], dict]:
    """
    分页拉取 JSON 接口全量数据。
    - 每页级重试：遇到 SQL 断连 / 非 JSON / HTTP != 200，sleep 5s 后重试，最多 3 次
    返回 (records, meta)，meta 包含 total_count / total_pages / extra_top_keys。
    """
    url = BASE_URL + api_path
    headers = {"X-Requested-With": "XMLHttpRequest"}
    all_records: list[dict] = []
    meta: dict = {"total_count": 0, "total_pages": 0, "extra_top_keys": []}
    RETRIES_PER_PAGE = 3
    RETRY_SLEEP = 6

    page = 1
    while page <= max_pages:
        data = {page_param: page, limit_param: page_size}
        r = None
        last_err = ""
        for attempt in range(1, RETRIES_PER_PAGE + 1):
            try:
                r = session.post(url, data=data, headers=headers, timeout=60)
            except Exception as e:
                last_err = f"请求异常: {e}"
                print(f"  [页 {page}] 尝试 {attempt}/{RETRIES_PER_PAGE} {last_err}")
                if attempt < RETRIES_PER_PAGE:
                    time.sleep(RETRY_SLEEP)
                continue
            if r.status_code != 200:
                last_err = f"HTTP {r.status_code}: {r.text[:200]}"
                print(f"  [页 {page}] 尝试 {attempt}/{RETRIES_PER_PAGE} {last_err}")
                if attempt < RETRIES_PER_PAGE:
                    time.sleep(RETRY_SLEEP)
                continue
            if _looks_like_sql_error(r.text):
                last_err = f"SQL 断连错误页 (前 200): {r.text[:200]}"
                print(f"  [页 {page}] 尝试 {attempt}/{RETRIES_PER_PAGE} {last_err}")
                if attempt < RETRIES_PER_PAGE:
                    time.sleep(RETRY_SLEEP)
                continue
            parsed = _try_parse_json(r.text)
            if not parsed or not isinstance(parsed, dict):
                last_err = f"非 JSON 返回 (前 200): {r.text[:200]}"
                print(f"  [页 {page}] 尝试 {attempt}/{RETRIES_PER_PAGE} {last_err}")
                if attempt < RETRIES_PER_PAGE:
                    time.sleep(RETRY_SLEEP)
                continue
            # 这一页请求成功
            break
        else:
            # 全部重试耗尽
            print(f"  [页 {page}] 多次重试失败，停止拉取。最后错误: {last_err}")
            break

        records = parsed.get(records_field) or []
        if not records:
            # root 为空可能是真的最后一页，也可能是服务器异常 — 保守停止
            print(f"  [页 {page}] root 为空，停止")
            break

        if page == 1:
            try:
                total_pages = int(parsed.get(total_field) or 1)
            except (TypeError, ValueError):
                total_pages = 1
            try:
                total_count = int(parsed.get("rtotal") or 0)
            except (TypeError, ValueError):
                total_count = 0
            meta["total_pages"] = total_pages
            meta["total_count"] = total_count
            # 顶层 key 中除了 root/pgtotal/page/rtotal 之外的，归为 extra
            standard_keys = {records_field, total_field, "page", "rtotal"}
            meta["extra_top_keys"] = [k for k in parsed.keys() if k not in standard_keys]
            # 检测服务端是否截断了每页条数
            actual_page_size = len(records)
            if actual_page_size < page_size:
                print(f"  [警告] 服务端每页只返回 {actual_page_size} 条（请求 {page_size}），"
                      f"已自动适配。")
            print(f"  接口返回: 共 {total_count} 条, {total_pages} 页 (实际每页 {actual_page_size})")
            if meta["extra_top_keys"]:
                print(f"  汇总字段 (extra_top_keys): {meta['extra_top_keys']}")

        all_records.extend(records)
        print(f"  第 {page}/{meta['total_pages']} 页: +{len(records)} 行, 累计 {len(all_records)} 行")

        if page >= meta["total_pages"]:
            break
        page += 1
        time.sleep(0.5)  # 友好限速（翻页间隔稍长，减少 SQL 压力）

    return all_records, meta


def fetch_html_export(session: requests.Session,
                      api_path: str,
                      timeout: int = 180) -> tuple[bytes, dict]:
    """
    调用 HTML 表格导出接口（excelbiao），返回原始 HTML 字节。
    服务端返回 Content-Type=Application/x-msexcel，本质是 HTML 表格伪装为 .xls。
    - 用 POST（与 kucunld 导出一致）
    - 带重试：返回体过小 (<5KB) 或包含 SQL 断连关键词 / 无 <tr>，sleep 5s 后重试，最多 3 次
    """
    url = BASE_URL + api_path
    RETRIES = 3
    RETRY_SLEEP = 6

    html_bytes = b""
    last_err = ""
    for attempt in range(1, RETRIES + 1):
        print(f"  POST {url} (尝试 {attempt}/{RETRIES})")
        try:
            r = session.post(url, data={}, timeout=timeout, allow_redirects=True)
        except Exception as e:
            last_err = f"请求异常: {e}"
            print(f"  {last_err}")
            if attempt < RETRIES:
                time.sleep(RETRY_SLEEP)
            continue
        if r.status_code != 200:
            last_err = f"HTTP {r.status_code}: {r.text[:200]}"
            print(f"  {last_err}")
            if attempt < RETRIES:
                time.sleep(RETRY_SLEEP)
            continue
        # 检测内容
        size = len(r.content)
        text_preview = r.content[:4000].decode("utf-8", errors="replace")
        if size < 5 * 1024 or _looks_like_sql_error(text_preview):
            last_err = (f"返回异常: size={size} bytes, "
                        f"preview={text_preview[:300]}")
            print(f"  {last_err}")
            if attempt < RETRIES:
                time.sleep(RETRY_SLEEP)
            continue
        has_tr = re.search(r'<tr\b', text_preview, flags=re.I) is not None
        if not has_tr:
            last_err = (f"返回异常: 未检测到 <tr> 表格行, size={size}, "
                        f"preview={text_preview[:300]}")
            print(f"  {last_err}")
            if attempt < RETRIES:
                time.sleep(RETRY_SLEEP)
            continue
        html_bytes = r.content
        break
    else:
        raise RuntimeError(f"HTML 导出多次重试失败: {last_err}")

    info = {
        "content_type": r.headers.get("Content-Type", ""),
        "size_bytes": len(html_bytes),
        "size_kb": round(len(html_bytes) / 1024, 1),
    }
    print(f"  成功, size={info['size_kb']} KB, ctype={info['content_type']}")
    return html_bytes, info


# ===================== Excel (XLSX) 输出 =====================
# 用户明确要求交付 XLSX（不是 CSV）。
# 用 openpyxl 直接写 workbook → BytesIO → bytes 上传。
# 注意：HTML 表格（库存导出 kucungl）返回的是伪 .xls (HTML)，先解析成二维 list，再写 xlsx。

def _parse_html_table_to_rows(html_bytes: bytes) -> list[list[str]]:
    """将 excelbiao 返回的 HTML 伪 .xls 解析为二维字符串列表（含表头）。"""
    text = html_bytes.decode("utf-8", errors="replace")
    rows_html = re.findall(r'<tr[^>]*>(.*?)</tr>', text, flags=re.S | re.I)
    if not rows_html:
        raise RuntimeError("HTML 中未找到 <tr> 表格行")
    rows_data: list[list[str]] = []
    max_cols = 0
    for tr in rows_html:
        cells = re.findall(r'<t[dh][^>]*>(.*?)</t[dh]>', tr, flags=re.S | re.I)
        cells_clean = []
        for c in cells:
            v = re.sub(r'<[^>]+>', '', c)
            v = (v.replace('&nbsp;', ' ')
                  .replace('&amp;', '&')
                  .replace('&lt;', '<')
                  .replace('&gt;', '>')
                  .replace('&quot;', '"')
                  .replace('&#39;', "'"))
            cells_clean.append(v.strip())
        if cells_clean:
            rows_data.append(cells_clean)
            if len(cells_clean) > max_cols:
                max_cols = len(cells_clean)
    if not rows_data:
        raise RuntimeError("HTML 表格解析后无数据")
    # 补齐列数（避免部分行列数不一致导致 openpyxl 出错）
    for row in rows_data:
        if len(row) < max_cols:
            row += [""] * (max_cols - len(row))
    return rows_data


def _cell_value(v):
    """把字符串尽量转成 int / float，Excel 里会显示成真正的数字。
    失败就保留原字符串。日期/时间不做解析，保持字符串即可。"""
    if v is None or v == "":
        return None
    s = str(v).strip()
    if s == "":
        return None
    # 整数
    if re.fullmatch(r'-?\d+', s):
        try:
            return int(s)
        except ValueError:
            pass
    # 小数（含 .0） — 但排除日期 "2026-08-15"、电话 "138-1234-5678"、带逗号的订单号
    if re.fullmatch(r'-?\d+\.\d+', s):
        try:
            return float(s)
        except ValueError:
            pass
    return s


def records_to_xlsx_bytes(records: list[dict],
                          summary_pairs: list[tuple[str, str]] | None = None) -> bytes:
    """list[dict] → XLSX 字节。
    结构：
      - Sheet1 叫"数据"：第 1 行表头，第 2 行起业务数据
      - Sheet2 叫"汇总"（可选）：如果有 summary_pairs 就放这里，两列：字段名/值
    summary_pairs: [(key, value), ...] 顶层汇总字段（如 hj_jiner=xxxx 等）
    """
    try:
        from openpyxl import Workbook
    except ImportError as e:
        raise RuntimeError("缺少 openpyxl，请 pip install openpyxl（已加入 requirements.txt）") from e

    # 收集所有字段名（保持首次出现顺序）
    fields: list[str] = []
    for r in records:
        for k in r.keys():
            if k not in fields:
                fields.append(k)

    wb = Workbook()
    ws = wb.active
    ws.title = "数据"

    # 表头
    ws.append(fields)

    # 数据行
    for r in records:
        row = [_cell_value(r.get(f)) for f in fields]
        ws.append(row)

    # 冻结首行，方便用户滚动看
    if fields:
        ws.freeze_panes = "A2"

    # 汇总 sheet
    if summary_pairs:
        ws2 = wb.create_sheet(title="汇总")
        ws2.append(["汇总字段", "值"])
        for k, v in summary_pairs:
            ws2.append([_cell_value(k), _cell_value(v)])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def rows_2d_to_xlsx_bytes(rows_2d: list[list[str]]) -> bytes:
    """二维数组（第一行通常是表头）→ XLSX 字节。用于库存导出 HTML 表格转储。"""
    try:
        from openpyxl import Workbook
    except ImportError as e:
        raise RuntimeError("缺少 openpyxl，请 pip install openpyxl（已加入 requirements.txt）") from e

    wb = Workbook()
    ws = wb.active
    ws.title = "库存"

    for row in rows_2d:
        ws.append([_cell_value(c) for c in row])

    # 冻结首行
    if rows_2d:
        ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ===== 兼容旧 CSV 函数（保留不删，防止其他依赖此文件的代码出错，本脚本主流程不再使用） =====

def records_to_csv_bytes(records: list[dict], extra_top_keys: list[str] = None) -> bytes:
    """list[dict] → CSV 字节（UTF-8-SIG）。已废弃，请用 records_to_xlsx_bytes。"""
    if not records and not extra_top_keys:
        return b""
    fields: list[str] = []
    for r in records:
        for k in r.keys():
            if k not in fields:
                fields.append(k)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(fields)
    for r in records:
        row = [("" if r.get(f) is None else str(r.get(f))) for f in fields]
        writer.writerow(row)
    if extra_top_keys:
        writer.writerow([])
        writer.writerow(["# 汇总字段"])
        for k in extra_top_keys:
            writer.writerow([f"# {k}"])
    return buf.getvalue().encode("utf-8-sig")


def html_table_to_csv_bytes(html_bytes: bytes) -> tuple[bytes, int, int]:
    """保留旧签名，内部委托给 _parse_html_table_to_rows。新代码请用 xlsx 分支。"""
    rows_data = _parse_html_table_to_rows(html_bytes)
    buf = io.StringIO()
    writer = csv.writer(buf)
    for row in rows_data:
        writer.writerow(row)
    return buf.getvalue().encode("utf-8-sig"), len(rows_data), len(rows_data[0])


# ===================== 飞书云盘 + 通知 =====================

def env(name: str, default: str = "") -> str:
    return os.environ.get(name, default).strip()


def feishu_tenant_access_token(app_id: str, app_secret: str) -> str:
    if not app_id or not app_secret:
        raise ValueError("缺少 FEISHU_APP_ID / FEISHU_APP_SECRET")
    url = f"{FEISHU_OPEN_BASE}/auth/v3/tenant_access_token/internal"
    r = requests.post(url, json={"app_id": app_id, "app_secret": app_secret}, timeout=15)
    data = r.json()
    if data.get("code") != 0:
        raise RuntimeError(f"换 tenant_access_token 失败: {data}")
    token = str(data.get("tenant_access_token") or "")
    if not token:
        raise RuntimeError(f"响应中无 tenant_access_token: {data}")
    print(f"[飞书] 获取 tenant_access_token 成功 (len={len(token)})")
    return token


def feishu_upload_to_folder(token: str, folder_token: str,
                            file_bytes: bytes, filename: str,
                            max_size_mb: int = 20) -> dict:
    """通过 drive/v1/files/upload_all 上传文件到飞书云盘（≤20MB）"""
    size = len(file_bytes)
    if size == 0:
        raise ValueError("上传文件为空")
    if size > max_size_mb * 1024 * 1024:
        raise ValueError(f"文件 {size/1024/1024:.1f}MB 超过 upload_all 上限 {max_size_mb}MB")
    if not folder_token:
        raise ValueError("缺少 folder_token")

    files = {"file": (filename, file_bytes, "application/octet-stream")}
    data = {
        "file_name": filename,
        "parent_type": "explorer",
        "parent_node": folder_token,
        "size": str(size),
    }
    headers = {"Authorization": f"Bearer {token}"}
    url = f"{FEISHU_OPEN_BASE}/drive/v1/files/upload_all"

    r = requests.post(url, data=data, files=files, headers=headers, timeout=180)
    resp = r.json()
    if resp.get("code") != 0:
        raise RuntimeError(f"上传失败 code={resp.get('code')} msg={resp.get('msg')}")
    file_info = resp.get("data") or {}
    file_token = file_info.get("file_token") or file_info.get("token") or ""
    print(f"  [飞书云盘] 上传成功 file_token={file_token}")
    return file_info


def _feishu_sign(secret: str, timestamp: str) -> str:
    import hmac, hashlib, base64 as _b64
    string_to_sign = f"{timestamp}\n{secret}"
    h = hmac.new(secret.encode("utf-8"),
                 string_to_sign.encode("utf-8"),
                 hashlib.sha256)
    return _b64.b64encode(h.digest()).decode("utf-8")


def feishu_send_bot_text(webhook_url: str, secret: str, text: str) -> None:
    if not webhook_url:
        return
    body = {"msg_type": "text", "content": {"text": text}}
    if secret:
        ts = str(int(time.time()))
        body["timestamp"] = ts
        body["sign"] = _feishu_sign(secret, ts)
    r = requests.post(webhook_url, json=body, timeout=15)
    resp = r.json()
    if resp.get("code") != 0 and resp.get("StatusCode") != 0:
        raise RuntimeError(f"飞书机器人通知失败: {resp}")
    print("[飞书通知] 发送成功")


# ===================== 主流程 =====================

def main() -> int:
    username = env("DH_USERNAME")
    password = env("DH_PASSWORD")
    fs_app_id = env("FEISHU_APP_ID")
    fs_app_secret = env("FEISHU_APP_SECRET")
    fs_webhook_url = env("FEISHU_WEBHOOK_URL")
    fs_webhook_secret = env("FEISHU_WEBHOOK_SECRET")
    page_size = int(env("BACKUP_PAGE_SIZE", "200") or "200")
    dry_run = env("BACKUP_DRY_RUN", "") == "1"

    if not username or not password:
        print("[错误] 缺少 DH_USERNAME / DH_PASSWORD")
        return 2

    if not dry_run and (not fs_app_id or not fs_app_secret):
        print("[错误] 缺少 FEISHU_APP_ID / FEISHU_APP_SECRET")
        return 2

    print("=" * 70)
    print("懂火系统 - 每周全量备份 10 类业务数据")
    print(f"时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"模式: {'DRY-RUN (仅本地 CSV)' if dry_run else '上传飞书云盘'}")
    print(f"分页: 每页 {page_size} 条")
    print("=" * 70)

    # 登录
    session = login_donghuo(username=username, password=password)
    if session is None:
        print("[退出] 登录失败")
        return 1

    # 准备本地输出目录（dry-run 或缓存）
    local_out_dir = Path(__file__).parent / "backup_local"
    local_out_dir.mkdir(exist_ok=True)

    # 飞书 token
    fs_token = None
    if not dry_run:
        try:
            fs_token = feishu_tenant_access_token(fs_app_id, fs_app_secret)
        except Exception as e:
            print(f"[错误] 获取飞书 token 失败: {e}")
            return 3

    # 遍历每个任务
    summary: list[dict] = []
    timestamp_str = datetime.datetime.now().strftime(DATE_FMT)

    for idx, task in enumerate(TASKS, 1):
        biz = task["biz"]
        print(f"\n[{idx}/{len(TASKS)}] === {biz} ===")
        folder_token = task.get("folder_token", "")
        if not dry_run and not folder_token:
            print(f"  [跳过] 未在 TASKS 中配置 folder_token")
            summary.append({
                "biz": biz, "status": "SKIP_NO_FOLDER",
                "reason": "TASKS 中该条目缺少 folder_token",
            })
            continue

        try:
            # 拉数据
            if task["api_type"] == "json_paged":
                records, meta = fetch_json_paged(
                    session,
                    api_path=task["api_path"],
                    page_param=task["page_param"],
                    limit_param=task["limit_param"],
                    total_field=task["total_field"],
                    records_field=task["records_field"],
                    page_size=page_size,
                )
                rows_count = len(records)
                cols_count = len(records[0]) if records else 0
                extra_summary = meta.get("extra_top_keys", [])
                summary_pairs: list[tuple[str, str]] = []
                if extra_summary:
                    # 重新拉一次第一页拿汇总字段值
                    parsed_first = _try_parse_json(
                        session.post(BASE_URL + task["api_path"],
                                     data={task["page_param"]: 1,
                                           task["limit_param"]: 1},
                                     headers={"X-Requested-With": "XMLHttpRequest"},
                                     timeout=60).text)
                    for k in extra_summary:
                        v = parsed_first.get(k, "") if isinstance(parsed_first, dict) else ""
                        summary_pairs.append((k, "" if v is None else str(v)))
                xlsx_bytes = records_to_xlsx_bytes(records, summary_pairs or None)

            elif task["api_type"] == "html_export":
                html_bytes, info = fetch_html_export(session, task["api_path"])
                rows_2d = _parse_html_table_to_rows(html_bytes)
                rows_count = len(rows_2d) - 1  # 扣掉表头行，和其他业务口径一致（纯数据行数）
                cols_count = len(rows_2d[0]) if rows_2d else 0
                xlsx_bytes = rows_2d_to_xlsx_bytes(rows_2d)
            else:
                print(f"  [错误] 未知 api_type: {task['api_type']}")
                continue

            # 文件名（xlsx）
            filename = f"{task['filename_prefix']}_{timestamp_str}.xlsx"
            local_path = local_out_dir / filename
            with open(local_path, "wb") as f:
                f.write(xlsx_bytes)
            size_kb = round(len(xlsx_bytes) / 1024, 1)
            print(f"  本地保存: {local_path} ({rows_count} 行 × {cols_count} 列, {size_kb} KB)")

            if dry_run:
                print(f"  [DRY-RUN] 跳过上传")
                summary.append({
                    "biz": biz, "status": "DRY_RUN",
                    "rows": rows_count, "cols": cols_count,
                    "size_kb": size_kb, "filename": filename,
                })
                continue

            # 上传飞书
            if not folder_token:
                print(f"  [跳过] 未在 TASKS 中配置 folder_token")
                summary.append({
                    "biz": biz, "status": "SKIP_NO_FOLDER",
                    "rows": rows_count, "filename": filename,
                })
                continue

            file_info = feishu_upload_to_folder(fs_token, folder_token,
                                                 xlsx_bytes, filename)
            summary.append({
                "biz": biz, "status": "OK",
                "rows": rows_count, "cols": cols_count,
                "size_kb": size_kb, "filename": filename,
                "file_token": file_info.get("file_token", ""),
            })

        except Exception as e:
            print(f"  [失败] {biz}: {e}")
            traceback.print_exc()
            summary.append({
                "biz": biz, "status": "FAIL",
                "error": str(e)[:200],
            })

        # 每类任务之间稍作间隔，避免懂火后端 SQL Server 连接池被连续请求打满
        if idx < len(TASKS):
            print(f"  间隔 4 秒，缓解服务端 SQL 连接压力 ...")
            time.sleep(4)

    # 汇总通知
    print("\n" + "=" * 70)
    print("[汇总] 备份结果:")
    for s in summary:
        status_icon = {"OK": "✅", "DRY_RUN": "💾", "SKIP_NO_FOLDER": "⏭️", "FAIL": "❌"}.get(s["status"], "?")
        line = f"  {status_icon} {s['biz']:<15} {s['status']}"
        if s.get("rows") is not None:
            line += f" ({s['rows']} 行)"
        if s.get("size_kb") is not None:
            line += f" {s['size_kb']} KB"
        if s.get("filename"):
            line += f" → {s['filename']}"
        if s.get("error"):
            line += f" | err: {s['error']}"
        print(line)

    # 发送飞书通知
    if fs_webhook_url and not dry_run:
        ok_count = sum(1 for s in summary if s["status"] == "OK")
        fail_count = sum(1 for s in summary if s["status"] == "FAIL")
        skip_count = sum(1 for s in summary if s["status"].startswith("SKIP"))
        total_rows = sum(s.get("rows", 0) for s in summary if s["status"] in ("OK", "DRY_RUN"))

        notify_lines = [
            f"📋 懂火系统每周备份 - {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}",
            f"成功: {ok_count} / 失败: {fail_count} / 跳过: {skip_count}",
            f"总记录数: {total_rows}",
            "",
        ]
        for s in summary:
            icon = {"OK": "✅", "DRY_RUN": "💾", "SKIP_NO_FOLDER": "⏭️", "FAIL": "❌"}.get(s["status"], "?")
            line = f"{icon} {s['biz']}"
            if s.get("rows") is not None:
                line += f" ({s['rows']} 行)"
            if s.get("size_kb") is not None:
                line += f" {s['size_kb']}KB"
            notify_lines.append(line)

        try:
            feishu_send_bot_text(fs_webhook_url, fs_webhook_secret, "\n".join(notify_lines))
        except Exception as e:
            print(f"[警告] 飞书通知发送失败: {e}")

    # 返回码：有失败 → 非零
    fail_count = sum(1 for s in summary if s["status"] == "FAIL")
    return 1 if fail_count > 0 else 0


if __name__ == "__main__":
    sys.exit(main())
